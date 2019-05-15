'''
# Created by:
# Selina Chua
# selina.a.chua@gmail.com
#
# This file contains the main code of the scraper.
# What it does:
#    1. Scrapes a given link for PDFs and downloads them.
#    2. Scraped the PDF for information and places info into a
#       spreadsheet.
'''

from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import resolve1

from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import requests, datetime, re, tabula, sys, csv
import camelot

from constants import *
from service import Service
from oldpdf_classes import oldPdfInfo
from newpdf_class import NewHospInfo, AmbulanceInfo, NewPDFInfo
from policy import WebPolicy
from criteria import Criteria
from shutil import rmtree
from os import makedirs


def scrape_all_pdfs(pdf_dict, sheet, pol_type):
    '''
    This function is the main pdf_scraper function. It iterates 
    through all the pdfs inside the dictionary and scrapes all
    information into the passed in excel sheet.
    '''
    wb = load_workbook(sheet)
    ws = wb.active
    input_row = 2

    for fund in pdf_dict:
        print(f"Scraping pdfs for {fund}")
        pdf_num = 1
        for link in pdf_dict[fund]:
            pdf_dest = TEMP_PDF + fund + str(pdf_num) + ".pdf"
            download_pdf(link, pdf_dest)
            print(f"Scraping PDF #{pdf_num} out of {len(pdf_dict[fund])}", end='\r')
            # Scrape the pdf for required information.
            pdf_info = scrape_pdf(pdf_dest, pol_type, pdf_dict[fund][link])
            # Populate the excel spreadsheet.
            if pdf_info.pdf_type == OLD_PDF:
                old_populate_excel(pdf_dict[fund][link], pdf_info, ws, input_row)
            elif pdf_info.pdf_type == NEW_PDF:
                new_populate_excel(pdf_dict[fund][link], pdf_info, ws, input_row)
            input_row += 1
            pdf_num += 1
        print("\n")
    wb.save(sheet)

    # Clean up temp folder and recreate empty one.
    rmtree(f"{sys.path[0]}/temp")
    makedirs(f"{sys.path[0]}/temp")


def scrape_pdf(pdf_dest, pol_type, web_policy):
    '''
    This function scrapes the pdf according to its type.
    '''
    pdf_text = pdf_to_text(pdf_dest)
    pdf_type = get_pdf_type(pdf_text)
    
    if pdf_type == NEW_PDF:
        pdf_info = scrape_new_pdf(pdf_dest, web_policy)
    elif pdf_type == OLD_PDF:
        pdf_info = scrape_old_pdf(pdf_dest, pdf_text, pol_type)
    else:
        pdf_info = "Couldn't identify PDF type."

    return pdf_info


def scrape_new_pdf(pdf_path, web_policy):
    '''
    This function scrapes information from the new PDF.
    '''
    # Gets number of pages in pdf. 
    file = open(pdf_path, 'rb')
    parser = PDFParser(file)
    document = PDFDocument(parser)
    n_pages = resolve1(document.catalog['Pages'])['Count']
    text = pdf_to_text(pdf_path)
    # Reads the pdf.
    ambulance_info = read_ambulance_page(text)
    if web_policy.criteria.pol_type == AMBULANCE_ONLY:
        return NewPDFInfo(NEW_PDF, None, {}, ambulance_info)

    hosp_info = read_hosp_page_new_pdf(text)
    gen_services = read_general_new_pdf(pdf_path, web_policy.criteria.pol_type, n_pages)

    return NewPDFInfo(NEW_PDF, hosp_info, gen_services, ambulance_info)


def new_populate_excel(web_policy, pdf_info, ws, input_row):
    '''
    This function populates excel from a new pdf class.
    '''
    # Input web policy information.
    if pdf_info.pdf_type == OLD_PDF:
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "OLD"
    elif pdf_info.pdf_type == NEW_PDF:
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "NEW"

    ws.cell(row=input_row, column=COL_POL_NAME).value = web_policy.name
    ws.cell(row=input_row, column=COL_FUND_NAME).value = web_policy.fund_name
    ws.cell(row=input_row, column=COL_PDF_LINK).value = web_policy.pdf_link
    ws.cell(row=input_row, column=COL_STATUS).value = web_policy.status
    ws.cell(row=input_row, column=COL_MOPREM).value = web_policy.premium 
    ws.cell(row=input_row, column=COL_EXCESS).value = web_policy.excess
    ws.cell(row=input_row, column=COL_COPAYMENT).value = web_policy.co_pay
    ws.cell(row=input_row, column=COL_AGE_DISC).value = web_policy.age_disc
    ws.cell(row=input_row, column=COL_MEDICARE).value = web_policy.medicare
    ws.cell(row=input_row, column=COL_POL_ID).value = web_policy.id
    ws.cell(row=input_row, column=COL_STATE).value = states[web_policy.criteria.state]
    ws.cell(row=input_row, column=COL_ADULTS).value = n_adults[web_policy.criteria.adults]
    ws.cell(row=input_row, column=COL_DPNDNTS).value = dpndnts[web_policy.criteria.dpndnts]
    ws.cell(row=input_row, column=COL_POL_TYPE).value = policy[web_policy.criteria.pol_type]
    ws.cell(row=input_row, column=COL_AVAIL).value = avail[web_policy.criteria.status]
    ws.cell(row=input_row, column=COL_CORP).value = corp[web_policy.criteria.corp]
    ws.cell(row=input_row, column=COL_AMBULANCE_EMER).value = pdf_info.ambulance.emer
    ws.cell(row=input_row, column=COL_AMBULANCE_FEE).value = pdf_info.ambulance.callout
    ws.cell(row=input_row, column=COL_AMBULANCE_OTHER).value = pdf_info.ambulance.other

    if web_policy.criteria.pol_type != AMBULANCE_ONLY:
        ws.cell(row=input_row, column=COL_TRAV_ACCOM_BEN).value = pdf_info.hosp.travel_ben
        ws.cell(row=input_row, column=COL_ACCIDENT_COV).value = pdf_info.hosp.accident_cover
        ws.cell(row=input_row, column=COL_PROV_ARR).value = pdf_info.hosp.prov_arr
        ws.cell(row=input_row, column=COL_ISSUE_DATE).value = pdf_info.hosp.issue_date
        ws.cell(row=input_row, column=COL_AVAIL_FOR).value = pdf_info.hosp.avail_for
        ws.cell(row=input_row, column=COL_WAIT_PERIODS).value = pdf_info.hosp.wait
        ws.cell(row=input_row, column=COL_OTHER).value = pdf_info.hosp.general_other
        ws.cell(row=input_row, column=COL_OTHER_HOSP).value = pdf_info.hosp.hosp_other
    
     # Inputting hospital cover details.
    covered = ""
    for c in web_policy.covered:
        covered += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_COVERED).value = covered
    not_covered = ""
    for c in web_policy.not_covered:
        not_covered += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_NOT_COVERED).value = not_covered
    limited_cover = ""
    for c in web_policy.limited_cover:
        limited_cover += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_LIMITED).value = limited_cover

    # Inputting general details.
    for service in pdf_info.services:
        col = ''
        s = service.lower()
        if 'general' in s:
            col = COL_GENERAL_DENTAL
        elif 'major' in s:
            col = COL_MAJOR_DENTAL
        elif 'endodontic' in s:
            col = COL_ENDODONTIC
        elif 'orthodontic' in s:
            col = COL_ORTHODONTIC
        elif 'optical' in s:
            col = COL_OPTICAL
        elif 'non pbs' in s:
            col = COL_NONPSBPHARM
        elif 'exercise physiology' in s:
            col = COL_EXERCISE_PHYSIO
        elif 'physio' in s:
            col = COL_PHYSIO
        elif 'chiro' in s:
            col = COL_CHIRO
        elif 'podiatry' in s:
            col = COL_PODIATRY
        elif 'psychology' in s:
            col = COL_PSYCH
        elif 'acupuncture' in s:
            col = COL_ACUPUNC
        elif 'naturopathy' in s:
            col = COL_NATUR
        elif 'massage' in s:
            col = COL_MASSAGE
        elif 'hearing aids' in s:
            col = COL_HEARING
        elif 'glucose' in s:
            col = COL_BLOOD
        elif 'audiology' in s:
            col = COL_AUDIO
        elif 'antenatal' in s:
            col = COL_NATAL
        elif 'chinese' in s:
            col = COL_CHINESE
        elif 'dietetics' in s:
            col = COL_DIETARY
        elif 'orthoptics' in s:
            col = COL_EYE_THERAPY
        elif 'health management' in s:
            col = COL_HEALTH_LIFE
        elif 'nursing' in s:
            col = COL_HOME_NURSING
        elif 'occupational therapy' in s:
            col = COL_OCCUPATIONAL_THER
        elif 'orthotics' in s:
            col = COL_ORTHOTICS
        elif 'osteopathy' in s:
            col = COL_OSTEOPATHY
        elif 'speech' in s:
            col = COL_SPEECH
        elif 'vaccinations' in s:
            col = COL_VACCINATIONS
        
        if col != '':
            ws.cell(row=input_row, column=col).value = pdf_info.services[service].wait 
            ws.cell(row=input_row, column=col+1).value = pdf_info.services[service].limits 
            ws.cell(row=input_row, column=col+2).value = pdf_info.services[service].max_benefits


def read_ambulance_page(pdf_text):
    '''
    Reads ambulance information.
    '''
    # Get emergency
    pattern = re.compile(r'Emergency: (.+)+')
    matches = pattern.finditer(pdf_text)
    emer = "-"
    for match in matches:
        emer = match.group(1)

    # Get callout fees
    pattern = re.compile(r'Call-out fees: (.+)+')
    matches = pattern.finditer(pdf_text)
    callout = "-"
    for match in matches:
        callout = match.group(1)
    
    # Get other
    other = "-"
    pattern = re.compile(r'Other features of this ambulance cover([.\s\S]+)For further')
    matches = pattern.finditer(pdf_text)
    for match in matches:
        other = match.group(1).strip()
        if "PolicyID:" in other:
            cut = other.index('Page')
            other = other[cut+12:].strip()
    
    return AmbulanceInfo(emer, callout, other)


def read_hosp_page_new_pdf(pdf_text):
    '''
    Gets:
    - DONE Waiting periods
    - DONE Availability 
    - DONE Available for 
    - DONE Travel and accom benefits 
    - DONE Accident Cover 
    - DONE Provider Arrangemnets
    - DONE Issue Date
    - DONE General Other
    - Hospital Other
    '''
    # Gets hospital other
    hosp_other = "-"
    pattern = re.compile(r'Other features of this hospital cover\s{2}((.+\s)+)')
    matches = pattern.finditer(pdf_text)
    for match in matches:
        hosp_other = match.group(1)

    # Gets issue date
    issue_date = "-"
    pattern = re.compile(r'Date statement issued: (.+)+')
    matches = pattern.finditer(pdf_text)
    for match in matches:
        string = pdf_text[match.span()[0]:match.span()[1]]
        issue_date = string

    # Get general other
    general_other = "-"
    # pattern = re.compile(r'Other features of this general treatment cover(\s{2}((.+\s)+)+)')
    pattern = re.compile(r'Other features of this general treatment cover([.\s\S]+)Ambulance cover')
    matches = pattern.finditer(pdf_text)
    for match in matches:
        general_other = match.group(1).strip()
        if "PolicyID:" in general_other:
            cut = general_other.index('Page')
            general_other = general_other[cut+12:len(general_other)-16]

    # Gets availability
    avail = "-"
    pattern = re.compile(r'\bAvailable in (.+)+')
    matches = pattern.finditer(pdf_text)
    for match in matches:
        string = pdf_text[match.span()[0]:match.span()[1]]
        avail = string

    # Gets travel benefits and accident cover.
    pattern = re.compile(r'\bThis policy (.+\s)+')
    matches = pattern.finditer(pdf_text)
    accident_cover = "-"
    travel_benefits = "-"
    for match in matches:
        string = pdf_text[match.span(0)[0]:match.span(0)[1]]
        if 'travel' in string:
            travel_benefits = string
        if 'accident' in string:
            accident_cover = string
    
    # Gets waiting period.
    pattern = re.compile(r'\bWaiting periods:\s{2}(.+\s)+')
    matches = pattern.finditer(pdf_text)
    wait = "-"
    for match in matches:
        string = pdf_text[match.span(0)[0]:match.span(0)[1]]
        wait = string

    # Gets provider arrangments
    pattern = re.compile(r'\bGeneral Treatment Cover\s{2}((.+\s)+)')
    matches = pattern.finditer(pdf_text)
    prov_arr = "-"
    for match in matches:
        string = pdf_text[match.span(0)[0]:match.span(0)[1]]
        if 'providers' in string:
            prov_arr = match.group(1)

    # Get available for.
    pattern = re.compile(r'\bMembership of this (.+\s)+')
    matches = pattern.finditer(pdf_text)
    avail_for = "-"
    for match in matches:
        string = pdf_text[match.span(0)[0]:match.span(0)[1]]
        avail_for = string
    
    return NewHospInfo(wait, avail, avail_for, travel_benefits, accident_cover, \
        prov_arr, issue_date, general_other, hosp_other)
    

def read_general_new_pdf(pdf_path, pol_type, n_pages):
    '''
    Gets all the general services.
    '''
    if pol_type == HOSPITAL:
        return {}

    if pol_type == GEN_TREATMENT:
        if n_pages == 1:
            pages = '1'
        else:
            pages = '1,2'
    else:
        if n_pages == 2:
            pages = '2'
        else: 
            pages = '2,3'
    # Gets the general tables.
    tables = camelot.read_pdf(pdf_path, pages)
    n_tables = len(tables)
    tables[0].to_csv(TEMP_CSV)
    
    services = {}
    f1 = open(TEMP_CSV)
    csv_f = csv.reader(f1)
    for row in csv_f:
        if 'Treatment' in row[SERVICE]:
            continue
        name = row[SERVICE]
        wait = row[NEW_WAIT_PERIOD]
        limits = row[NEW_BENEFIT_LIMITS]
        if limits == "":
            limits = "Same as previous."
        max_ben = row[NEW_MAX_BENEFITS]

        serv = Service(name, "Yes", wait, limits, max_ben)
        services[name] = serv
    f1.close()

    if n_tables > 1:
        tables[n_tables-1].to_csv(TEMP_CSV)
        f2 = open(TEMP_CSV)
        csv_f = csv.reader(f2)
        for row in csv_f:
            name = row[SERVICE]
            wait = row[NEW_WAIT_PERIOD]
            limit = row[NEW_BENEFIT_LIMITS]
            max_ben = row[NEW_MAX_BENEFITS]
            serv = Service(name, "Yes", wait, limit, max_ben)
            services[name] = serv
        f2.close()
    return services
    

def old_populate_excel(web_policy, pdf_info, ws, input_row):
    '''
    This function populates old PDF into excel sheets.
    '''
    # Input web policy information.
    if pdf_info.pdf_type == OLD_PDF:
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "OLD"
    elif pdf_info.pdf_type == NEW_PDF:
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "NEW"

    ws.cell(row=input_row, column=COL_POL_NAME).value = web_policy.name
    ws.cell(row=input_row, column=COL_FUND_NAME).value = web_policy.fund_name
    ws.cell(row=input_row, column=COL_PDF_LINK).value = web_policy.pdf_link
    ws.cell(row=input_row, column=COL_STATUS).value = web_policy.status
    ws.cell(row=input_row, column=COL_MOPREM).value = web_policy.premium 
    ws.cell(row=input_row, column=COL_EXCESS).value = web_policy.excess
    ws.cell(row=input_row, column=COL_COPAYMENT).value = web_policy.co_pay
    ws.cell(row=input_row, column=COL_AGE_DISC).value = web_policy.age_disc
    ws.cell(row=input_row, column=COL_MEDICARE).value = web_policy.medicare
    ws.cell(row=input_row, column=COL_POL_ID).value = web_policy.id
    ws.cell(row=input_row, column=COL_STATE).value = states[web_policy.criteria.state]
    ws.cell(row=input_row, column=COL_ADULTS).value = n_adults[web_policy.criteria.adults]
    ws.cell(row=input_row, column=COL_DPNDNTS).value = dpndnts[web_policy.criteria.dpndnts]
    ws.cell(row=input_row, column=COL_POL_TYPE).value = policy[web_policy.criteria.pol_type]
    ws.cell(row=input_row, column=COL_AVAIL).value = avail[web_policy.criteria.status]
    ws.cell(row=input_row, column=COL_CORP).value = corp[web_policy.criteria.corp]
    ws.cell(row=input_row, column=COL_PROV_ARR).value = pdf_info.prov_arr
    ws.cell(row=input_row, column=COL_ISSUE_DATE).value = pdf_info.issue_date
    ws.cell(row=input_row, column=COL_AVAIL_FOR).value = pdf_info.avail_for
    ws.cell(row=input_row, column=COL_WAIT_PERIODS).value = pdf_info.wait
    ws.cell(row=input_row, column=COL_OTHER).value = pdf_info.general_other
    ws.cell(row=input_row, column=COL_TRAV_ACCOM_BEN).value = web_policy.hosp_accom

     # Inputting hospital cover details.
    covered = ""
    for c in web_policy.covered:
        covered += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_COVERED).value = covered
    not_covered = ""
    for c in web_policy.not_covered:
        not_covered += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_NOT_COVERED).value = not_covered
    limited_cover = ""
    for c in web_policy.limited_cover:
        limited_cover += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_LIMITED).value = limited_cover
    ws.cell(row=input_row, column=COL_OTHER_HOSP).value = pdf_info.hosp_other

    # Inputting general details.
    if pdf_info.services is None:
        return 
    for service in pdf_info.services:
        col = ''
        s = service.lower()
        if 'general' in s:
            col = COL_GENERAL_DENTAL
        elif 'major' in s:
            col = COL_MAJOR_DENTAL
        elif 'endodontic' in s:
            col = COL_ENDODONTIC
        elif 'orthodontic' in s:
            col = COL_ORTHODONTIC
        elif 'optical' in s:
            col = COL_OPTICAL
        elif 'non pbs' in s:
            col = COL_NONPSBPHARM
        elif 'exercise physiology' in s:
            col = COL_EXERCISE_PHYSIO
        elif 'physio' in s:
            col = COL_PHYSIO
        elif 'chiro' in s:
            col = COL_CHIRO
        elif 'podiatry' in s:
            col = COL_PODIATRY
        elif 'psychology' in s:
            col = COL_PSYCH
        elif 'acupuncture' in s:
            col = COL_ACUPUNC
        elif 'naturopathy' in s:
            col = COL_NATUR
        elif 'massage' in s:
            col = COL_MASSAGE
        elif 'hearing aids' in s:
            col = COL_HEARING
        elif 'glucose' in s:
            col = COL_BLOOD
        elif 'audiology' in s:
            col = COL_AUDIO
        elif 'antenatal' in s:
            col = COL_NATAL
        elif 'chinese' in s:
            col = COL_CHINESE
        elif 'dietetics' in s:
            col = COL_DIETARY
        elif 'orthoptics' in s:
            col = COL_EYE_THERAPY
        elif 'health management' in s:
            col = COL_HEALTH_LIFE
        elif 'nursing' in s:
            col = COL_HOME_NURSING
        elif 'occupational therapy' in s:
            col = COL_OCCUPATIONAL_THER
        elif 'orthotics' in s:
            col = COL_ORTHOTICS
        elif 'osteopathy' in s:
            col = COL_OSTEOPATHY
        elif 'speech' in s:
            col = COL_SPEECH
        elif 'vaccinations' in s:
            col = COL_VACCINATIONS
        elif 'ambulance' in s and pdf_info.pdf_type == OLD_PDF:
            col = COL_AMBULANCE_WP
        
        if col != '':
            ws.cell(row=input_row, column=col).value = pdf_info.services[service].wait 
            ws.cell(row=input_row, column=col+1).value = pdf_info.services[service].limits 
            ws.cell(row=input_row, column=col+2).value = pdf_info.services[service].max_benefits


def scrape_old_pdf(pdf_path, pdf_text, pol_type):
    '''
    This function scrapes the old pdf for wanted information.
    '''

    pdf_hosp_info = read_hosp_page_old_pdf(pdf_path, pdf_text, pol_type)
    if pol_type == HOSPITAL:
        return pdf_hosp_info
    pdf_info = read_general_old_pdf(pdf_path, pdf_text, pdf_hosp_info, pol_type)
    return pdf_info
    

def read_general_old_pdf(pdf_path, pdf_text, oldpdf_class, pol_type):
    '''
    This functions reads the general treatment page in the old pdf.
    '''
    # Calculates page of the general.
    page = 2
    if pol_type == GEN_TREATMENT:
        page = 1

    # Converts the general table in PDF into CSV format. 
    tabula.convert_into(pdf_path, TEMP_CSV, \
        lattice=True, spreadsheet=True, pages=page, output_format='csv')

    f = open(TEMP_CSV)
    csv_f = csv.reader(f)

    prov_arr = "None"
    services = {}
    other = "-"
    for row in csv_f:
        # Special cases:
        if 'PROVIDER ARRANGEMENTS:' in row[SERVICE]:
            prov_arr = row[SERVICE]
        elif 'FEATURES' in row[SERVICE]:
            other = row[SERVICE]
        elif 'SERVICES' in row[SERVICE] \
            or not row[SERVICE]:
            continue
        # Get the general treatment stuff. 
        else:
            name = row[SERVICE]
            wait = row[WAITING_PERIOD]
            # If there is no waiting period, service is not covered.
            if '-' in wait:
                cover = "No"
            else:
                cover = "Yes"
            limit = row[BENEFIT_LIMITS]
            max_benefits = row[MAX_BENEFITS]
            # If max benefits is empty, then it is in limits and limit is the same as previous.
            if not max_benefits:
                max_benefits = limit 
                limit = '-'
            if cover == "Yes" and limit == '-':
                limit = "Same as previous."
            cur_service = Service(name, cover, wait, limit, max_benefits)
            services[name] = cur_service
    f.close()
    oldpdf_class.prov_arr = prov_arr
    oldpdf_class.services = services
    oldpdf_class.general_other = other
    return oldpdf_class


def read_hosp_page_old_pdf(pdf_path, pdf_text, pol_type):
    '''
    This function reads the hospital page in the pdf and scrapes 
    for its issue date and available for.
    '''
    # Grabs issue date.
    try:
        issue_date = str(re.search(r'issued (.*)\n', pdf_text).group(1))
    except:
        issue_date = "Can't find issue date"
    # Grabs available for information.
    first_lines = pdf_text.rsplit('\n')
    avail_for = ""
    for line in first_lines:
        if 'Residents' in line:
            avail_for = line
            break
    
    if pol_type == GEN_TREATMENT:
        return oldPdfInfo(OLD_PDF, "-", {}, issue_date, avail_for, "-", "-", "-", "-")
    # Read rest of hospital page.
    tabula.convert_into(pdf_path, TEMP_CSV, \
        lattice=True, spreadsheet=True, pages=1, output_format='csv')

    f = open(TEMP_CSV)
    csv_f = csv.reader(f)

    waiting_period = "-"; payable = "-"; hosp_other = "-"
    for row in csv_f:
        # Find waiting period.
        if 'HOW LONG ARE THE WAITING' in row[SERVICE]:
            waiting_period = row[INFO]
        # Find hospital payables. 
        if 'WILL I HAVE TO PAY' in row[SERVICE]:
            payable = row[INFO]
        # Other hospital features
        if 'OTHER FEATURES' in row[SERVICE]:
            hosp_other = row[INFO]
    f.close()

    return oldPdfInfo(OLD_PDF, "-", {}, issue_date, avail_for, payable, waiting_period, "-", hosp_other)
    

def download_pdf(url, dest):
    '''
    This function downloads the pdf given a url
    and saves it to the given destination.
    '''
    pdf = requests.get(url)
    with open(dest, 'wb') as f:
        f.write(pdf.content)


def create_excel(destination):
    '''
    Sets up Excel Category Sheet & Bold title given sheetname
    '''
    wb = Workbook()
    ws = wb.active
    colname = ["PDF Type", "Name", "Fund", "PDFLink", "Status", "Excess", \
            "Monthly Premium", "State", "Adults", "Dependants", \
            "Availability", "Policy Type", "Corporate Product", \
            "Hospital Cover During Visit", "Hospital Services not Covered", \
            "Hospital Services Limited Cover", "Waiting periods", "Copayment", \
            "Other Hospital Cover Features", "General Dental - WP", \
            "General Dental - Limits", "General Dental - Max Benefits", \
            "Major Dental - WP", "Major Dental - Limits", "Major Dental - Max Benefits", \
            "Endodontic - WP", "Endodontic - Limits", "Endodontic - Max Benefits", \
            "Orthodontic - WP", "Orthodontic - Limits", "Orthodontic - Max Benefits", \
            "Optical - WP", "Optical - Limits", "Optical - Max Benefits", \
            "NonPBSPharmaceuticals - WP", "NonPBSPharmaceuticals - Limits", \
            "NonPBSPharmaceuticals - Max Benefits", "Physio - WP", "Physio - Limits", \
            "Physio - Max Benefits", "Chiropractic - WP", "Chiropractic - Limits", \
            "Chiropractic - Max Benefits", "Podiatry - WP", "Podiatry - Limits", \
            "Podiatry - Max Benefits", "Psychology - WP", "Psychology - Limits", \
            "Psychology - Max Benefits", "Acupuncture - WP", "Acupuncture - Limits", \
            "Acupuncture - Max Benefits", "Naturopathy - WP", "Naturopathy - Limits", \
            "Naturopathy - Max Benefits", "Massage - WP", "Massage - Limits", \
            "Massage - Max Benefits", "HearingAids - WP", "HearingAids - Limits", \
            "HearingAids - Max Benefits", "BloodGlucose Monitoring - WP", \
            "BloodGlucose Monitoring - Limits", "BloodGlucose Monitoring - Max Benefits", \
            "Audiology - WP", "Audiology - Limits", "Audiology - Max Benefits", \
            "Ante-natal/Post-natal - WP", "Ante-natal/Post-natal - Limits", "Ante-natal/Post-natal - Max Benefits", \
            "Chinese Medicine - WP", "Chinese Medicine - Limits", "Chinese Medicine - Max Benefits", \
            "Dietary Advice - WP", "Dietary Advice - Limits", "Dietary Advice - Max Benefits", \
            "Exercise Physiology - WP", "Exercise Physiology - Limits", "Audiology - Max Benefits", \
            "Eye Therapy - Emergency", "Eye Therapy - Call out fees", "Eye Therapy - other information", \
            "Health Management - WP", "Health Management - Limits", "Health Management - Max Benefits", \
            "Home nursing - WP", "Home nursing - Limits", "Home nursing - Max Benefits", \
            "Occupational therapy - WP", "Occupational therapy - Limits", "Occupational therapy - Max Benefits", \
            "Orthotics - WP", "Orthotics - Limits", "Orthotics - Max Benefits", \
            "Osteopathy - WP", "Osteopathy - Limits", "Osteopathy - Max Benefits", \
            "Speech Therapy - WP", "Speech Therapy - Limits", "Speech Therapy - Max Benefits", \
            "Vaccinations - WP", "Vaccinations - Limits", "Vaccinations - Max Benefits", \
            "Ambulance - Emergency", "Ambulance - Call Out Fees", "Ambulance - Other", \
            "Other Treatment Cover Features", "Medicare Surcharge Levy", "Issue Date", \
            "Available for", "Provider Arrangements", "Youth discount", \
            "Travel and accommodation beneft", "Policy ID", "Accident cover"]

    #Creates Bold Column Titles
    for i in range(len(colname)):
        ws.cell(row = 1, column = i + 1).value = colname[i]
        ws.cell(row = 1, column = i + 1).font = Font(size = 14, bold = True)
    wb.save(filename=destination)


def pdf_to_text(path):
    '''
    Converts the pdf into text. Taken from online.
    Source can be found in scraperConst.py as
    PDF_TO_TEXT_SRC.
    '''
    manager = PDFResourceManager()
    retstr = BytesIO()
    layout = LAParams(all_texts=True)
    device = TextConverter(manager, retstr, laparams=layout)
    filepath = open(path, 'rb')
    interpreter = PDFPageInterpreter(manager, device)

    for page in PDFPage.get_pages(filepath, check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    filepath.close()
    device.close()
    retstr.close()
    return text.decode('utf-8')


def get_pdf_type(pdf_text):
    '''
    There are 2 types of PDF: the old, and the new.
    This function returns OLD_PDF for old ones, and 
    NEW_PDF for new ones.
    '''
    if OLD_TITLE in pdf_text:
        return OLD_PDF
    elif NEW_TITLE in pdf_text:
        return NEW_PDF


if __name__ == "__main__":
    pdf_path = f"{sys.path[0]}/temp/temp.pdfNIB11.pdf"
    pol_type = GEN_TREATMENT
    n_pages  = 2
    text = pdf_to_text(pdf_path)
    # print(text)
    crit = Criteria("000101")
    web_policy = WebPolicy("ESH", "Gold Hospital", "link", "Open", crit, "100", "none", "No copay", 'No agedisc', 'exempted', 'a', 'b', ['a','b'], ['c','d'], ['e','f'], 'ABC')
    pdf_info = scrape_new_pdf(pdf_path, web_policy)
    # pdf_info = scrape_old_pdf(pdf_path, text, web_policy.criteria.pol_type)
    # print(pdf_info.hosp.general_other)
    # # read_hosp_page_new_pdf(text)
    # crit = Criteria("000200")
    # web_policy = WebPolicy("ACA", "Something", "abc", "Open", crit, "100", "no excess", "no copay", "no age disc", "no medicare", "no", "no", ['a', 'b'], ['c','d'], ['e','f'], "other", "J100")
    # pdf_info = scrape_new_pdf(pdf_path, web_policy)

    # sheet = "results.xlsx"
    # create_excel(sheet)
    # wb = load_workbook(sheet)
    # ws = wb.active
    # print("inputting into sheet")
    # new_populate_excel(web_policy, pdf_info, ws, 2)
    # wb.save(sheet)

    # text = pdf_to_text(r"temp/NJKD20.pdf")
    # print(text)
    # # scrape_new_pdf(text)
    # crit = Criteria("000100")
    # web_policy = WebPolicy("ACA", "Gold Deluxe Hospital", "link", "Open", crit, "100", "no excess", "no copay", "no age disc", "No medicare", "No", "no", ["a", 'b'], ['c', 'd'], ['e', 'f'], "other", "J20")
    # print(web_policy.criteria.pol_type)
    # pdf_info = scrape_old_pdf(text, web_policy.criteria.pol_type)

    # # Creates excel sheet.
    # line = "000000"
    # sheet = "results.xlsx"
    # create_excel(sheet)
    # wb = load_workbook(sheet)
    # ws = wb.active
    # old_populate_excel(web_policy, pdf_info, ws, 2)
    # wb.save(sheet)